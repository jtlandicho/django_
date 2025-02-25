from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.http import HttpResponse, StreamingHttpResponse, HttpResponseForbidden
from django.utils import timezone
from django.db.models import Avg, Count, Sum
from datetime import datetime, timedelta, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl.utils
from openpyxl.cell.cell import MergedCell
from .models import ParkingSlot, ParkingHistory
from pymongo import MongoClient
from django.views.decorators.http import require_GET
from django.core.cache import caches
import json
import time
import logging

# Custom decorator to check for admin access
def is_admin(user):
    return user.is_superuser

from datetime import datetime


from django.shortcuts import redirect

def disagree(request):
    # Redirect users to a different page if they disagree
    return redirect("")  # Change "home" to the appropriate URL name

def HomePage(request):
    from .models import ParkingSlot
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Get all - slots
    car_slots = ParkingSlot.objects.filter(vehicle_type='car')
    motorcycle_slots = ParkingSlot.objects.filter(vehicle_type='motorcycle')
    
    # Log detailed slot information
    logger.warning("Car Slots Debug:")
    for slot in car_slots:
        logger.warning(f"Slot {slot.slot_number}: Status = {slot.status}, Raw Status = {slot.status}")
        # Check if status is exactly matching the choices
        logger.warning(f"Status Matches Choices: {slot.status in dict(ParkingSlot.STATUS_CHOICES).keys()}")
    
    logger.warning("Motorcycle Slots Debug:")
    for slot in motorcycle_slots:
        logger.warning(f"Slot {slot.slot_number}: Status = {slot.status}, Raw Status = {slot.status}")
        # Check if status is exactly matching the choices
        logger.warning(f"Status Matches Choices: {slot.status in dict(ParkingSlot.STATUS_CHOICES).keys()}")
    
    # Calculate statistics with dynamic total based on availability
    car_stats = {
        'total': car_slots.filter(status='available').count(),  
        'available': car_slots.filter(status='available').count(),
        'occupied': car_slots.filter(status='occupied').count(),
        'maintenance': car_slots.filter(status='maintenance').count(),
        'slots': car_slots
    }
    
    motorcycle_stats = {
        'total': motorcycle_slots.filter(status='available').count(),  
        'available': motorcycle_slots.filter(status='available').count(),
        'occupied': motorcycle_slots.filter(status='occupied').count(),
        'maintenance': motorcycle_slots.filter(status='maintenance').count(),
        'slots': motorcycle_slots
    }
    
    # Log statistics and raw queryset
    logger.warning(f"Car Slots Raw Queryset: {list(car_slots.values('slot_number', 'status'))}")
    logger.warning(f"Motorcycle Slots Raw Queryset: {list(motorcycle_slots.values('slot_number', 'status'))}")
    logger.warning(f"Car Stats: {car_stats}")
    logger.warning(f"Motorcycle Stats: {motorcycle_stats}")
    
    current_time = datetime.now().strftime('%I:%M:%S %p')
    
    return render(request, 'home.html', {
        'user': request.user,
        'car_stats': car_stats,
        'motorcycle_stats': motorcycle_stats,
        'current_time': current_time
    })

def SignUpPage(request):
    if request.method == 'POST':
        uname = request.POST.get("username")
        email = request.POST.get("email")
        pass1 = request.POST.get('password1')
        pass2 = request.POST.get('password2')

        # Check if passwords match
        if pass1 != pass2:
            messages.error(request, "Passwords do not match")
            return redirect('signup')

        # Check for existing username or email
        if User.objects.filter(username=uname).exists():
            messages.error(request, "Username already exists.")
            return redirect('signup')
        
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already registered.")
            return redirect('signup')

        # Create the user
        my_user = User.objects.create_user(username=uname, email=email, password=pass1)
        my_user.save()

        messages.success(request, "Registration successful! Please log in.")
        return redirect('login')
    
    return render(request, 'signup.html')

def LoginPage(request):
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        pass1 = request.POST.get('pass')
        user = authenticate(request, username=username, password=pass1)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, "Username or Password is incorrect!")
            return redirect('login')

    return render(request, 'login.html')

def LogoutPage(request):
    logout(request)
    return redirect('login')

def tools(request):
    if request.method == 'POST':
        tool = request.POST.get('tool')
        
        if tool == 'slot_management':
            action = request.POST.get('action')
            slots = request.POST.getlist('slots')
            
            if not slots:
                messages.error(request, 'Please select at least one parking slot.')
                return redirect('tools')
                
            slots_qs = ParkingSlot.objects.filter(id__in=slots)
            
            if action == 'reset':
                slots_qs.update(status='available', is_occupied=False)
                messages.success(request, f'Successfully reset {len(slots)} parking slots.')
                
            elif action == 'maintenance':
                # Toggle maintenance mode
                for slot in slots_qs:
                    slot.status = 'available' if slot.status == 'maintenance' else 'maintenance'
                    slot.save()
                messages.success(request, f'Updated maintenance mode for {len(slots)} slots.')
                
            elif action == 'type':
                new_type = request.POST.get('new_type')
                if new_type in ['car', 'motorcycle']:
                    slots_qs.update(vehicle_type=new_type)
                    messages.success(request, f'Updated vehicle type for {len(slots)} slots.')
                    
        elif tool == 'data_management':
            action = request.POST.get('action')
            
            if action == 'cleanup':
                try:
                    from django.core.management import call_command
                    retention_days = int(request.POST.get('retention', 30))
                    deleted_count = call_command('cleanup_old_records', days=retention_days)
                    
                    if deleted_count >= 0:
                        messages.success(request, f'Cleaned up {deleted_count} old parking records.')
                    else:
                        messages.error(request, 'Failed to clean up records. Check server logs for details.')
                except Exception as e:
                    messages.error(request, f'Error cleaning up records: {str(e)}')
                return redirect('tools')
                
            elif action == 'backup':
                # Create database backup
                from django.core import serializers
                from itertools import chain
                
                timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                filename = f'parksense_backup_{timestamp}.json'
                
                response = HttpResponse(content_type='application/json')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                
                # Serialize all relevant models
                models = [ParkingSlot, ParkingHistory]
                objects = chain(*[model.objects.all() for model in models])
                
                serializers.serialize('json', objects, stream=response)
                return response
                
            elif action == 'reset_counters':
                try:
                    from django.core.management import call_command
                    result = call_command('reset_analytics')
                    messages.success(request, result)
                except Exception as e:
                    messages.error(request, f'Error resetting analytics: {str(e)}')
                return redirect('tools')
    
    # Get the current year and month
    now = timezone.now()
    current_year = now.year
    current_month = now.month
    
    # Get the earliest record's year
    earliest_record = ParkingHistory.objects.order_by('timestamp').first()
    start_year = earliest_record.timestamp.year if earliest_record else current_year
    
    # Generate list of years from earliest to current
    available_years = list(range(start_year, current_year + 1))
    
    # Generate list of months with names
    months = [
        {'number': 1, 'name': 'January'},
        {'number': 2, 'name': 'February'},
        {'number': 3, 'name': 'March'},
        {'number': 4, 'name': 'April'},
        {'number': 5, 'name': 'May'},
        {'number': 6, 'name': 'June'},
        {'number': 7, 'name': 'July'},
        {'number': 8, 'name': 'August'},
        {'number': 9, 'name': 'September'},
        {'number': 10, 'name': 'October'},
        {'number': 11, 'name': 'November'},
        {'number': 12, 'name': 'December'}
    ]
    
    return render(request, 'tools.html', {
        'available_years': available_years,
        'months': months,
        'current_year': current_year,
        'current_month': current_month,
        'parking_slots': ParkingSlot.objects.all().order_by('slot_number')
    })
    return render(request, 'analytics.html', {
        'available_years': available_years,
        'months': months,
        'current_year': current_year,
        'current_month': current_month,
        'parking_slots': ParkingSlot.objects.all().order_by('slot_number')
    })

from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET

@csrf_exempt
@require_GET
def public_slots_updates(request):
    """
    Server-Sent Events (SSE) endpoint specifically for the public slots display
    Streams only the count of available slots
    """
    def event_stream():
        sse_logger.info("Public Slots SSE Stream Initialization")
        
        # Send headers for SSE
        response = StreamingHttpResponse(content_type='text/event-stream')
        response['Cache-Control'] = 'no-cache'
        response['X-Accel-Buffering'] = 'no'
        
        while True:
            try:
                # Get the Redis cache
                cache = caches['default']
                
                # Try to get parking data from cache
                parking_data = cache.get('parking_slots_status')
                
                if parking_data:
                    # Parse the cached JSON data
                    try:
                        if isinstance(parking_data, str):
                            parking_stats = json.loads(parking_data)
                        else:
                            parking_stats = parking_data
                            
                        available_car_slots = sum(1 for slot in parking_stats.values() 
                                                if slot.get('vehicle_type') == 'car' and 
                                                slot.get('status') == 'available')
                        available_motorcycle_slots = sum(1 for slot in parking_stats.values() 
                                                       if slot.get('vehicle_type') == 'motorcycle' and 
                                                       slot.get('status') == 'available')
                    except (json.JSONDecodeError, KeyError) as e:
                        sse_logger.error(f"Error parsing cache data: {e}")
                        # Fallback to database
                        car_slots = ParkingSlot.objects.filter(vehicle_type='car')
                        motorcycle_slots = ParkingSlot.objects.filter(vehicle_type='motorcycle')
                        available_car_slots = car_slots.filter(status='available').count()
                        available_motorcycle_slots = motorcycle_slots.filter(status='available').count()
                else:
                    # Fallback to database if cache is empty
                    car_slots = ParkingSlot.objects.filter(vehicle_type='car')
                    motorcycle_slots = ParkingSlot.objects.filter(vehicle_type='motorcycle')
                    available_car_slots = car_slots.filter(status='available').count()
                    available_motorcycle_slots = motorcycle_slots.filter(status='available').count()
                
                # Format the data for the public display
                data = {
                    'available_car_slots': available_car_slots,
                    'available_motorcycle_slots': available_motorcycle_slots
                }
                
                # Send initial data immediately
                yield f"data: {json.dumps(data)}\n\n"
                
                # Flush the response to ensure data is sent
                if hasattr(response, 'flush'):
                    response.flush()
                    
                time.sleep(1)  # Check for updates every second
                
            except Exception as e:
                sse_logger.error(f"Error in public slots SSE stream: {str(e)}")
                time.sleep(1)  # Wait before retrying
    
    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response

@require_GET
def parking_slots(request):
    # Get the Redis cache
    cache = caches['default']
    
    # Try to get parking data from cache
    parking_data = cache.get('parking_slots_status')
    
    if parking_data:
        # Parse the cached JSON data
        try:
            parking_stats = json.loads(parking_data)
            available_car_slots = sum(1 for slot in parking_stats.values() 
                                    if slot.get('vehicle_type') == 'car' and slot.get('status') == 'available')
            available_motorcycle_slots = sum(1 for slot in parking_stats.values() 
                                           if slot.get('vehicle_type') == 'motorcycle' and slot.get('status') == 'available')
        except (json.JSONDecodeError, KeyError):
            # Fallback to database if cache data is invalid
            car_slots = ParkingSlot.objects.filter(vehicle_type='car')
            motorcycle_slots = ParkingSlot.objects.filter(vehicle_type='motorcycle')
            available_car_slots = car_slots.filter(status='available').count()
            available_motorcycle_slots = motorcycle_slots.filter(status='available').count()
    else:
        # Fallback to database if cache is empty
        car_slots = ParkingSlot.objects.filter(vehicle_type='car')
        motorcycle_slots = ParkingSlot.objects.filter(vehicle_type='motorcycle')
        available_car_slots = car_slots.filter(status='available').count()
        available_motorcycle_slots = motorcycle_slots.filter(status='available').count()
    
    return render(request, 'slots.html', {
        'available_car_slots': available_car_slots,
        'available_motorcycle_slots': available_motorcycle_slots
    })

import logging
import json
import time
from django.http import StreamingHttpResponse
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required
from django.core.cache import caches
from django.conf import settings

# Configure a specific logger for SSE events
sse_logger = logging.getLogger('parking_sse')
sse_logger.setLevel(logging.INFO)

# If no handler exists, add a console handler
if not sse_logger.handlers:
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(logging.Formatter(
        '%(asctime)s - SSE - %(levelname)s: %(message)s'
    ))
    sse_logger.addHandler(console_handler)

from django.conf import settings

@require_GET
def parking_slot_updates(request):
    """
    Server-Sent Events (SSE) endpoint for real-time parking slot updates
    Streams live updates from cache, preventing direct database reads
    """
    def event_stream():
        """
        Generator function to create a continuous event stream of parking slot updates
        """
        sse_logger.info("SSE Stream Initialization")
        
        # Get the Redis cache
        cache = caches['default']
        
        while True:
            try:
                # Get latest data from cache or database
                parking_data = cache.get('parking_slots_status')
                if not parking_data:
                    # Fallback to database
                    car_slots = ParkingSlot.objects.filter(vehicle_type='car')
                    motorcycle_slots = ParkingSlot.objects.filter(vehicle_type='motorcycle')
                    
                    # Format data for SSE
                    formatted_data = {
                        'car_slots': [
                            {
                                'sensor_id': slot.sensor_id,
                                'status': slot.status,
                                'last_updated': slot.last_updated.isoformat() if slot.last_updated else None
                            } for slot in car_slots
                        ],
                        'motorcycle_slots': [
                            {
                                'sensor_id': slot.sensor_id,
                                'status': slot.status,
                                'last_updated': slot.last_updated.isoformat() if slot.last_updated else None
                            } for slot in motorcycle_slots
                        ]
                    }
                else:
                    # Parse cached data and format it
                    try:
                        if isinstance(parking_data, str):
                            parking_data = json.loads(parking_data)
                        
                        # Convert flat structure to grouped structure
                        car_slots = []
                        motorcycle_slots = []
                        
                        for slot_id, slot_data in parking_data.items():
                            slot_info = {
                                'sensor_id': slot_id,
                                'status': slot_data['status'],
                                'last_updated': slot_data.get('last_updated')
                            }
                            if slot_data['vehicle_type'] == 'car':
                                car_slots.append(slot_info)
                            else:
                                motorcycle_slots.append(slot_info)
                        
                        formatted_data = {
                            'car_slots': car_slots,
                            'motorcycle_slots': motorcycle_slots
                        }
                    except (json.JSONDecodeError, KeyError) as e:
                        sse_logger.error(f"Error parsing cached data: {e}")
                        continue
                
                # Send the formatted data
                sse_logger.info("Sending parking slot update")
                yield f"data: {json.dumps(formatted_data)}\n\n"
                
                time.sleep(1)  # Check for updates every second
                
            except Exception as e:
                sse_logger.error(f"Error in SSE stream: {str(e)}")
                time.sleep(1)  # Wait before retrying
                
                # Sleep to control update frequency and prevent tight looping
                time.sleep(settings.SSE_UPDATE_INTERVAL)
            
            except GeneratorExit:
                # Handle generator being closed
                sse_logger.info("SSE Stream Closed")
                break
            except Exception as e:
                sse_logger.error(f"SSE Stream Error: {e}", exc_info=True)
                yield f"data: {json.dumps({'error': str(e)})}\n\n"
                time.sleep(settings.SSE_UPDATE_INTERVAL)  # Wait before retrying
    
    # Create streaming response
    response = StreamingHttpResponse(
        event_stream(), 
        content_type='text/event-stream; charset=utf-8'
    )
    
    # Set SSE-specific headers
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    
    # Add CORS headers if needed
    response['Access-Control-Allow-Origin'] = '*'
    
    sse_logger.info(f"SSE Response Created")
    
    return response

def analytics(request):
    import json
    
    # Get all parking slots
    parking_slots = ParkingSlot.objects.all()
    
    # Calculate current statistics
    total_slots = parking_slots.count()
    occupied_slots = parking_slots.filter(status='Occupied').count()
    occupancy_rate = (occupied_slots / total_slots * 100) if total_slots > 0 else 0
    
    # Calculate real 24h utilization for each slot
    now = timezone.now()
    day_ago = now - timedelta(days=1)
    
    for slot in parking_slots:
        # Count occupied time in last 24h for this slot
        occupied_records = ParkingHistory.objects.filter(
            slot_id=slot.id,
            status='Occupied',
            timestamp__gte=day_ago,
            timestamp__lte=now
        )
        
        total_occupied_time = timedelta()
        for record in occupied_records:
            # Use the duration field directly
            if record.duration:
                total_occupied_time += record.duration
        
        # Add utilization rate directly to the slot object
        slot.real_time_utilization = (total_occupied_time.total_seconds() / (24 * 3600)) * 100
    
    # Calculate average duration and format durations
    avg_duration = timedelta()
    total_count = 0
    for slot in parking_slots:
        if slot.total_occupancy_count > 0:
            avg_duration += slot.total_occupancy_time
            total_count += slot.total_occupancy_count
        # Format the slot's average duration
        if slot.total_occupancy_count > 0:
            seconds = slot.average_occupancy_duration.total_seconds()
            hours = int(seconds // 3600)
            minutes = int((seconds % 3600) // 60)
            slot.formatted_duration = f"{hours}h {minutes}m" if hours > 0 else f"{minutes}m"
        else:
            slot.formatted_duration = "No data"
            
    if total_count > 0:
        avg_duration = avg_duration / total_count
        avg_duration_str = f"{int(avg_duration.total_seconds() // 3600)}h {int((avg_duration.total_seconds() % 3600) // 60)}m"
    else:
        avg_duration_str = "No data"
    
    # Calculate peak hours (last 24h)
    peak_slots = parking_slots.filter(status='Occupied').count()
    peak_time = timezone.now().strftime('%I:%M %p')
    
    # Prepare chart data
    # 1. Occupancy Trend (24h)
    now = timezone.now()
    hours = [(now - timedelta(hours=x)).strftime('%I %p') for x in range(23, -1, -1)]
    
    # Get real occupancy data for the last 24 hours
    occupancy_values = []
    for hour in range(23, -1, -1):
        start_time = now - timedelta(hours=hour+1)
        end_time = now - timedelta(hours=hour)
        
        # Get all records that overlap with this hour
        records = ParkingHistory.objects.filter(
            timestamp__lt=end_time,
            status='Occupied'
        )
        
        # Calculate total occupied time in this hour
        total_occupied_seconds = 0
        for record in records:
            if record.duration:
                # Convert duration to total occupied time in this hour
                record_start = record.timestamp
                record_end = record.timestamp + record.duration
                
                # Only count the overlap with current hour
                if record_end > start_time:
                    overlap_start = max(record_start, start_time)
                    overlap_end = min(record_end, end_time)
                    overlap_duration = overlap_end - overlap_start
                    total_occupied_seconds += overlap_duration.total_seconds()
        
        # Calculate occupancy rate for this hour (as percentage of total slot-hours)
        total_slot_seconds = total_slots * 3600  # total available seconds (slots * 1 hour)
        hourly_rate = (total_occupied_seconds / total_slot_seconds * 100) if total_slot_seconds > 0 else 0
        occupancy_values.append(round(hourly_rate, 1))
    
    occupancy_data = {
        'labels': hours,
        'values': occupancy_values
    }
    
    # 2. Vehicle Distribution
    vehicle_counts = parking_slots.values('vehicle_type').annotate(count=Count('id'))
    vehicle_data = {
        'labels': [dict(ParkingSlot.VEHICLE_TYPES)[vt['vehicle_type']] for vt in vehicle_counts],
        'values': [vt['count'] for vt in vehicle_counts]
    }
    
    # 3. Slot Utilization (using real-time calculated values)
    utilization_data = {
        'labels': [f"Slot {slot.slot_number}" for slot in parking_slots],
        'values': [round(slot.real_time_utilization, 1) for slot in parking_slots],
        'colors': ['#3498db' if slot.vehicle_type == 'car' else '#e74c3c' for slot in parking_slots]
    }
    
    # Handle Excel export
    if request.GET.get('export') == 'excel':
        report_type = request.GET.get('report_type', 'daily')
        
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active

        # Get the date range based on report type
        today = timezone.now()
        if report_type == 'daily':
            date_str = request.GET.get('date')
            if date_str:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            else:
                target_date = today.date()
            start_date = datetime.combine(target_date, time.min)
            end_date = datetime.combine(target_date, time.max)
            ws.title = f"Daily Report {target_date}"
            filename = f'ParkSense_Daily_Report_{target_date}.xlsx'
        
        elif report_type == 'monthly':
            month_str = request.GET.get('month')
            if month_str:
                year, month = map(int, month_str.split('-'))
            else:
                year, month = today.year, today.month
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(seconds=1)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(seconds=1)
            ws.title = f"Monthly Report {year}-{month:02d}"
            filename = f'ParkSense_Monthly_Report_{year}_{month:02d}.xlsx'
        
        else:  # yearly
            year = int(request.GET.get('year', today.year))
            start_date = datetime(year, 1, 1)
            end_date = datetime(year, 12, 31, 23, 59, 59)
            ws.title = f"Yearly Report {year}"
            filename = f'ParkSense_Yearly_Report_{year}.xlsx'

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        centered = Alignment(horizontal="center")

        # Get historical data for the period
        history_data = ParkingHistory.objects.filter(
            timestamp__range=(start_date, end_date)
        )

        # Summary Section
        ws['A1'] = f"Parking Analytics Summary ({start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = centered

        # Calculate summary statistics
        total_slots = ParkingSlot.objects.count()
        avg_occupancy = history_data.aggregate(Avg('occupancy_rate'))['occupancy_rate__avg'] or 0
        avg_duration = history_data.aggregate(Avg('duration'))['duration__avg'] or timedelta()
        peak_data = history_data.order_by('-occupied_count').first()
        peak_time = peak_data.timestamp.strftime('%I:%M %p') if peak_data else 'No data'

        # Add summary data
        summary_data = [
            ['Total Slots', total_slots],
            ['Average Occupancy Rate', f"{avg_occupancy:.1f}%"],
            ['Average Duration', f"{avg_duration.total_seconds()/3600:.1f} hours"],
            ['Peak Time', peak_time]
        ]

        headers = ['Metric', 'Value']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered

        for row, (metric, value) in enumerate(summary_data, start=3):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=str(value))

        # Vehicle Distribution Section
        row = 8
        ws[f'A{row}'] = "Vehicle Distribution"
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].alignment = centered
        row += 1

        headers = ['Vehicle Type', 'Total Slots', 'Avg Available', 'Avg Occupied', 'Avg Occupancy Rate']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
        row += 1

        for vehicle_type in ['Car', 'Motorcycle']:
            slots = ParkingSlot.objects.filter(vehicle_type=vehicle_type.lower())
            total = slots.count()
            type_data = history_data.filter(slot__vehicle_type=vehicle_type.lower())
            avg_occupied = type_data.aggregate(Avg('occupied_count'))['occupied_count__avg'] or 0
            avg_available = total - avg_occupied
            avg_rate = type_data.aggregate(Avg('occupancy_rate'))['occupancy_rate__avg'] or 0

            ws.cell(row=row, column=1, value=vehicle_type)
            ws.cell(row=row, column=2, value=total)
            ws.cell(row=row, column=3, value=f"{avg_available:.1f}")
            ws.cell(row=row, column=4, value=f"{avg_occupied:.1f}")
            ws.cell(row=row, column=5, value=f"{avg_rate:.1f}%")
            row += 1

        # Detailed Stats Section
        row += 2
        ws[f'A{row}'] = "Detailed Slot Statistics"
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].alignment = centered
        row += 1

        headers = ['Slot Number', 'Vehicle Type', 'Total Hours', 'Total Occupancy Count', 'Avg Duration (hours)', 'Utilization Rate']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
        row += 1

        slots = ParkingSlot.objects.all().order_by('vehicle_type', 'slot_number')
        for slot in slots:
            slot_data = history_data.filter(slot=slot)
            total_duration = slot_data.aggregate(Sum('duration'))['duration__sum'] or timedelta()
            total_hours = total_duration.total_seconds() / 3600
            total_count = slot_data.count()
            avg_duration = total_hours / total_count if total_count > 0 else 0
            period_hours = (end_date - start_date).total_seconds() / 3600
            utilization_rate = (total_hours / period_hours * 100) if period_hours > 0 else 0

            ws.cell(row=row, column=1, value=f"{slot.vehicle_type} {slot.slot_number}")
            ws.cell(row=row, column=2, value=slot.get_vehicle_type_display())
            ws.cell(row=row, column=3, value=f"{total_hours:.1f}")
            ws.cell(row=row, column=4, value=total_count)
            ws.cell(row=row, column=5, value=f"{avg_duration:.1f}")
            ws.cell(row=row, column=6, value=f"{utilization_rate:.1f}%")
            row += 1

        # Adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value) if cell.value is not None else '') 
                        for cell in column_cells 
                        if not isinstance(cell, openpyxl.cell.cell.MergedCell))
            ws.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length + 2

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response
    
    context = {
        'total_slots': total_slots,
        'occupancy_rate': round(occupancy_rate, 1),
        'avg_duration': avg_duration_str,
        'peak_hours': peak_time,
        'parking_slots': parking_slots,
        'occupancy_data': json.dumps(occupancy_data),
        'vehicle_data': json.dumps(vehicle_data),
        'utilization_data': json.dumps(utilization_data),
        # No need to pass real_time_utilization since it's on each slot object
    }
    
    return render(request, 'analytics.html', context)
